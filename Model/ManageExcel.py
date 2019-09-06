import openpyxl
import csv
import sys
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Border, colors, PatternFill, Side
from Model import RuleEngine

ColumnNameSet = ['Nickname','Analyzer ID','Date','Time','Sample No.','Measurement Mode','Discrete','WBC(10^3/uL)','RBC(10^6/uL)','HGB(g/dL)','HCT(%)','IRF(%)','MCH(pg)','MCHC(g/dL)','MCV(fL)','PLT(10^3/uL)','RDW-SD(fL)','RET#(10^6/uL)','RET%(%)','IRF(%)'] #'Off Score (10*Hb(g/dL)-60*square root(ret%))'
ColumnLabelSet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S']
ColumnWideSet = [10,12,12,10,16,8,18,15,13,12,7,7,7,10,10,8,8,8,8,7,20]
MaxCol = 155

def MainExcel(DataDict):
    print('=== Main Excel ==============================================================')
    InputPathFile = DataDict['InputPathFile']
    OutputPathFile = DataDict['OutputPathFile']
    ThisWorkBook = CreateWorkBook() #Create new excel file
    
    ListExcel = ConvertCSVtoList(InputPathFile) #List of Excel row by row
    ConvertTextValue(ListExcel)
    HeaderSheet = ListExcel.pop(0) #get header first row of input
    DeleteExcessColumn(ListExcel) #delete unuse column
    RuleEngine.CreateOffScore(ListExcel) #calculate formular
    AppendToWorkbook(ThisWorkBook, ListExcel, HeaderSheet) #append remain row to new excel file

    
    #Design Excel
    Row = len(ListExcel)
    Col = 19 #Fix Value
    FixColumnWide(ThisWorkBook) #set width
    FillHeader(ThisWorkBook) #add fix text
    FixBorder(ThisWorkBook, Row, Col) #add border
    
    #Save WorkBook
    print("Save Excel File ...")
    ThisWorkBook.save(OutputPathFile)

def ConvertTextValue(ListExcel):
    for row in ListExcel:
        row[0] = RuleEngine.TextValue(row[0])
    return ListExcel

def AppendToWorkbook(ThisWorkBook,ListExcel, HeaderSheet):
    ThisWorkSheet = ThisWorkBook.active
    ListExcel.insert(0,HeaderSheet)
    #ListExcel[1].append("OffScore")
    for Row in ListExcel:
        ThisWorkSheet.append(Row)
    return ThisWorkBook

def DeleteExcessColumn(ListExcel):
    print('DeleteExcessColumn ...')
    Pos_False = MarkDeleteColumn(ListExcel)
    for Count,RowExcel in enumerate(ListExcel):
        for delColumn in reversed(Pos_False):
            del RowExcel[delColumn]
        
    print('DeleteExcessColum: Done')
    return ListExcel

def MarkDeleteColumn(ListExcel):
    print("MarkDeleteColumn ...")
    Pos_False = [];Pos_True = []
    ColumnNameSet_Input = ListExcel[0]
    
    for Position,ColumnName_Input in enumerate(ColumnNameSet_Input):
        if ColumnName_Input not in ColumnNameSet: #ColumnNameSet is column that user want to see
            Pos_False.append(Position)
        else:
            Pos_True.append(Position)
    print("MarkDeleteColumn: Done")
    return Pos_False

def FillHeader(ThisWorkBook):
    ThisWorkSheet = ThisWorkBook.active
    #Header Analyzer Value
    for idx in range(7,19):
        ThisWorkSheet.cell(2,idx).fill =StyleSheet("FillYellow")
    #Header Rule Engine
    ThisWorkSheet.cell(2,19).fill =StyleSheet("FillPurple")
    
def CreateMetrixExcel(ThisWorkBook):
    print("CreateMetrixExcel ...")
    ThisWorkSheet = ThisWorkBook.active
    for row in ThisWorkSheet.iter_rows():
        yield [cell.value for cell in row]
    print("CreateMetrixExcel: Done")

def CreateWorkBook():
    ThisWorkBook = openpyxl.Workbook()
    return ThisWorkBook
    
def ConvertCSVtoExcel(InputPathFile):
    print("Converting CSV to Excel")
    DelimeterChar = ','
    ThisWorkBook = openpyxl.Workbook()
    ThisWorkSheet = ThisWorkBook.active
    
    with open(InputPathFile) as f:
        reader = csv.reader(f, delimiter = DelimeterChar)
        for row in reader:
            ThisWorkSheet.append(row)
    return ThisWorkBook

def ConvertCSVtoList(InputPathFile):
    print("Converting CSV to Excel ...")
    DelimeterChar = ','
    ListExcel = []
    
    with open(InputPathFile) as f:
        reader = csv.reader(f, delimiter = DelimeterChar)
        for row in reader:
            ListExcel.append(row)
    print("Converting CSV to Excel: Done")
    return ListExcel

def FixColumnWide(ThisWorkBook):
    ThisWorkSheet = ThisWorkBook.active
    for idx,i in enumerate(ColumnLabelSet):
        ThisWorkSheet.column_dimensions[i].width = ColumnWideSet[idx]
    
    return 0

def FixBorder(ThisWorkBook,Row,Col):
    ThisWorkSheet = ThisWorkBook.active
    StrokeBorder(ThisWorkSheet, 1, Col, 2, Row)
    return ThisWorkSheet

def StrokeBorder(workSheet, min_X, max_X, min_Y, max_Y):
    min_X -= 1
    max_X -= 1
    min_Y -= 1
    max_Y -= 1
    BorderStyle = StyleSheet('BorderNor')
    for i, row in enumerate(workSheet):
        if min_Y <= i <= max_Y:
            for j, cell in enumerate(row):
                if min_X <= j <= max_X:
                    cell.border = BorderStyle
    return workSheet

def StyleSheet(key):
    StyleDict = {}
    StyleDict['AliignMid'] = Alignment(horizontal="center",vertical="center",wrap_text=True)
    StyleDict['FillGray'] = PatternFill(start_color='f2f2f2', fill_type='solid')
    StyleDict['FillPurple'] = PatternFill(start_color='ccd9ff', fill_type='solid')
    StyleDict['FillPink'] = PatternFill(start_color='F9B1B1', fill_type='solid')
    StyleDict['BorderNor'] = Border(top = Side(border_style='thin', color='FF000000'),    
                          right = Side(border_style='thin', color='FF000000'), 
                          bottom = Side(border_style='thin', color='FF000000'),
                          left = Side(border_style='thin', color='FF000000'))
    StyleDict['BorderNor_Left'] = Border(left = Side(border_style='thin', color='FF000000'))
    StyleDict['FillYellow'] = PatternFill(start_color='FFFF00', fill_type='solid')
    StyleDict['FontHead'] = Font(name = 'TH SarabunPSK', size = 16, bold = True)
    StyleDict['FontText'] = Font(name = 'TH SarabunPSK', size = 16, bold = False)
    return StyleDict[key]